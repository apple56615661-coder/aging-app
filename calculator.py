from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

AGING_WEIGHTS = {
    "외벽": 0.25,
    "창호": 0.25,
    "배관": 0.25,
    "설비": 0.25,
}

def is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")

def to_float(v):
    if is_blank(v):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("%", "")
    if s == "":
        return None
    try:
        return float(s)
    except:
        return None

def normalize_yes_no(v):
    if is_blank(v):
        return None
    s = str(v).strip().lower()
    yes_set = {"yes", "y", "예", "ㅇ", "true", "1"}
    no_set = {"no", "n", "아니오", "아니요", "x", "false", "0"}
    if s in yes_set:
        return True
    if s in no_set:
        return False
    return None

def score_to_zero_one(score):
    if score is None:
        return None
    s = max(1.0, min(4.0, float(score)))
    return (s - 1.0) / 3.0

def weighted_score_from_scores(data, weights):
    total = 0.0
    used_weight = 0.0
    for code, weight in weights.items():
        v = to_float(data.get(code))
        if v is None:
            continue
        total += score_to_zero_one(v) * weight
        used_weight += weight
    if used_weight == 0:
        return 0.0
    return total * (100.0 / used_weight)

def ratio_to_1to4_age(ratio):
    if ratio is None:
        return None
    if ratio < 0.50:
        return 1
    elif ratio < 0.80:
        return 2
    elif ratio < 1.20:
        return 3
    else:
        return 4

def percent_drop_to_1to4(drop_value):
    if drop_value is None:
        return None
    x = float(drop_value)
    if x > 1:
        x = x / 100.0
    x = max(0.0, x)
    if x < 0.10:
        return 1
    elif x < 0.20:
        return 2
    elif x < 0.40:
        return 3
    else:
        return 4

def repair_count_to_1to4(count_value):
    if count_value is None:
        return None
    c = int(round(float(count_value)))
    if c <= 0:
        return 1
    elif c == 1:
        return 2
    elif 2 <= c <= 3:
        return 3
    else:
        return 4

def repeat_count_to_1to4(count_value):
    if count_value is None:
        return None
    c = int(round(float(count_value)))
    if c <= 0:
        return 1
    elif c == 1:
        return 2
    elif c == 2:
        return 3
    else:
        return 4

def parse_building_sheet(ws):
    values = {}
    for row in range(1, ws.max_row + 1):
        code = ws[f"A{row}"].value
        val = ws[f"E{row}"].value
        if isinstance(code, str) and code.strip():
            values[code.strip()] = val
    return values

def calc_wall_aging(data):
    weights = {
        "wall_crack": 20,
        "wall_spalling": 20,
        "wall_rebar": 15,
        "wall_leak": 15,
        "wall_finish": 10,
        "wall_joint": 10,
        "wall_insulation": 10,
    }
    return weighted_score_from_scores(data, weights)

def calc_window_aging(data):
    weights = {
        "win_open": 15,
        "win_lock": 10,
        "win_leak": 15,
        "win_air": 15,
        "win_cond": 10,
        "win_thermal": 15,
        "win_frame": 10,
        "win_sealant": 5,
        "win_glass": 5,
    }
    return weighted_score_from_scores(data, weights)

def calc_pipe_aging(data):
    age_ratio = to_float(data.get("pipe_age_ratio"))
    age_score = ratio_to_1to4_age(age_ratio)
    part_a = score_to_zero_one(age_score) * 20 if age_score is not None else None

    corrosion = to_float(data.get("pipe_corrosion"))
    leak = to_float(data.get("pipe_leak"))

    part_b_sum = 0.0
    part_b_weight = 0.0
    if corrosion is not None:
        part_b_sum += score_to_zero_one(corrosion) * 15
        part_b_weight += 15
    if leak is not None:
        part_b_sum += score_to_zero_one(leak) * 10
        part_b_weight += 10
    part_b = part_b_sum * (25 / part_b_weight) if part_b_weight > 0 else None

    flow_drop_score = percent_drop_to_1to4(to_float(data.get("pipe_flow_drop")))
    pressure_drop_score = percent_drop_to_1to4(to_float(data.get("pipe_pressure_drop")))
    perf_scores = [x for x in [flow_drop_score, pressure_drop_score] if x is not None]
    if perf_scores:
        avg_perf = sum(perf_scores) / len(perf_scores)
        part_c = score_to_zero_one(avg_perf) * 20
    else:
        part_c = None

    water = to_float(data.get("pipe_water"))
    part_d = score_to_zero_one(water) * 15 if water is not None else None

    repair_count = repair_count_to_1to4(to_float(data.get("pipe_repair_count")))
    repeat_yes = normalize_yes_no(data.get("pipe_repeat_yes"))
    repair_final = None
    if repair_count is not None:
        repair_final = repair_count + (0.5 if repeat_yes else 0.0)
        repair_final = min(repair_final, 4.0)
    part_e = score_to_zero_one(repair_final) * 10 if repair_final is not None else None

    support = to_float(data.get("pipe_support"))
    part_f = score_to_zero_one(support) * 10 if support is not None else None

    parts = {"A": part_a, "B": part_b, "C": part_c, "D": part_d, "E": part_e, "F": part_f}
    section_weights = {"A": 20, "B": 25, "C": 20, "D": 15, "E": 10, "F": 10}

    used_weight = sum(section_weights[k] for k, v in parts.items() if v is not None)
    if used_weight == 0:
        return 0.0

    raw = sum(v for v in parts.values() if v is not None)
    return raw * (100.0 / used_weight)

def calc_equipment_aging(data):
    def yn_to_1or4(code):
        yn = normalize_yes_no(data.get(code))
        if yn is None:
            return None
        return 4 if yn else 1

    age_ratio = to_float(data.get("eq_age_ratio"))
    age_score = ratio_to_1to4_age(age_ratio)
    outer = to_float(data.get("eq_outer"))
    leak_score = yn_to_1or4("eq_leak_yes")

    part_a_list = []
    if age_score is not None:
        part_a_list.append(score_to_zero_one(age_score) * 10)
    if outer is not None:
        part_a_list.append(score_to_zero_one(outer) * 10)
    if leak_score is not None:
        part_a_list.append(score_to_zero_one(leak_score) * 10)
    part_a = sum(part_a_list) if part_a_list else None

    flow = to_float(data.get("eq_flow"))
    pressure = to_float(data.get("eq_pressure"))
    temp = to_float(data.get("eq_temp"))
    part_b_list = []
    if flow is not None:
        part_b_list.append(score_to_zero_one(flow) * 10)
    if pressure is not None:
        part_b_list.append(score_to_zero_one(pressure) * 10)
    if temp is not None:
        part_b_list.append(score_to_zero_one(temp) * 10)
    part_b = sum(part_b_list) if part_b_list else None

    control = to_float(data.get("eq_control"))
    safety_score = yn_to_1or4("eq_safety_yes")
    part_c_list = []
    if control is not None:
        part_c_list.append(score_to_zero_one(control) * 10)
    if safety_score is not None:
        part_c_list.append(score_to_zero_one(safety_score) * 10)
    part_c = sum(part_c_list) if part_c_list else None

    repair_score = repair_count_to_1to4(to_float(data.get("eq_repair_count")))
    repeat_score = repeat_count_to_1to4(to_float(data.get("eq_repeat_count")))
    abandon_score = yn_to_1or4("eq_abandon_yes")
    part_d_list = []
    if repair_score is not None:
        part_d_list.append(score_to_zero_one(repair_score) * 10)
    if repeat_score is not None:
        part_d_list.append(score_to_zero_one(repeat_score) * 5)
    if abandon_score is not None:
        part_d_list.append(score_to_zero_one(abandon_score) * 5)
    part_d = sum(part_d_list) if part_d_list else None

    parts = {"A": part_a, "B": part_b, "C": part_c, "D": part_d}
    section_weights = {"A": 30, "B": 30, "C": 20, "D": 20}

    used_weight = sum(section_weights[k] for k, v in parts.items() if v is not None)
    if used_weight == 0:
        return 0.0

    raw = sum(v for v in parts.values() if v is not None)
    return raw * (100.0 / used_weight)

def calc_wall_urgency(data):
    weights = {
        "wall_urg_crack": 20,
        "wall_urg_spalling": 20,
        "wall_urg_rebar": 20,
        "wall_urg_leak": 15,
        "wall_urg_falling": 25,
    }
    return weighted_score_from_scores(data, weights)

def calc_window_urgency(data):
    weights = {
        "win_urg_open": 20,
        "win_urg_lock": 15,
        "win_urg_leak": 20,
        "win_urg_frame": 20,
        "win_urg_glass": 25,
    }
    return weighted_score_from_scores(data, weights)

def calc_pipe_urgency(data):
    weights = {
        "pipe_urg_leak": 20,
        "pipe_urg_repeat": 15,
        "pipe_urg_risk": 20,
        "pipe_urg_corrosion": 10,
        "pipe_urg_flow": 10,
        "pipe_urg_pressure": 10,
        "pipe_urg_support": 15,
    }
    return weighted_score_from_scores(data, weights)

def calc_equipment_urgency(data):
    weights = {
        "eq_urg_leak": 15,
        "eq_urg_safety": 20,
        "eq_urg_repair": 15,
        "eq_urg_abandon": 10,
        "eq_urg_outer": 5,
        "eq_urg_flow": 10,
        "eq_urg_pressure": 10,
        "eq_urg_control": 15,
    }
    return weighted_score_from_scores(data, weights)

def evaluate_excel(file_path):
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    skip_names = {"📋사용안내", "사용 안내", "사용안내"}

    results = []

    for ws in wb.worksheets:
        if ws.title in skip_names:
            continue

        data = parse_building_sheet(ws)

        wall_aging = round(calc_wall_aging(data), 2)
        window_aging = round(calc_window_aging(data), 2)
        pipe_aging = round(calc_pipe_aging(data), 2)
        equipment_aging = round(calc_equipment_aging(data), 2)

        aging_total = round(
            wall_aging * AGING_WEIGHTS["외벽"] +
            window_aging * AGING_WEIGHTS["창호"] +
            pipe_aging * AGING_WEIGHTS["배관"] +
            equipment_aging * AGING_WEIGHTS["설비"],
            2
        )

        wall_urg = round(calc_wall_urgency(data), 2)
        window_urg = round(calc_window_urgency(data), 2)
        pipe_urg = round(calc_pipe_urgency(data), 2)
        equipment_urg = round(calc_equipment_urgency(data), 2)

        results.append({
            "건물명": ws.title,
            "외벽(100)": wall_aging,
            "창호(100)": window_aging,
            "배관(100)": pipe_aging,
            "설비(100)": equipment_aging,
            "노후도 종합(100)": aging_total,
            "외벽 긴급도(100)": wall_urg,
            "창호 긴급도(100)": window_urg,
            "배관 긴급도(100)": pipe_urg,
            "설비 긴급도(100)": equipment_urg,
        })

    result_df = pd.DataFrame(results)

    ranking_df = result_df[["건물명", "노후도 종합(100)"]].copy()
    ranking_df["노후도 순위"] = ranking_df["노후도 종합(100)"].rank(
        ascending=False, method="min"
    ).astype(int)
    ranking_df = ranking_df.sort_values(["노후도 순위", "건물명"]).reset_index(drop=True)

    urgency_df = result_df[[
        "건물명",
        "외벽 긴급도(100)",
        "창호 긴급도(100)",
        "배관 긴급도(100)",
        "설비 긴급도(100)"
    ]].copy()

    return result_df, ranking_df, urgency_df